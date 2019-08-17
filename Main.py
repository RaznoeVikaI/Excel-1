import openpyxl
wb = openpyxl.load_workbook('задача для кандидата.xlsx')
# Get a sheet by name
#sheet = wb.get_sheet_by_name('списки')
sheet = wb['списки']
#sheet.cell(row=1, column=2).value

#Название филиала
#print(sheet.cell(row=1, column=2).value)
#name_articles=sheet.cell(row=1, column=2).value
#print(name_articles)

#Запомним столбец "Название филиалов"
names_filials = []
# Print out values in column 2
for i in range(2, 5):
     #print(i, sheet.cell(row=i, column=2).value)
    names_filials.append(sheet.cell(row=i, column=2).value)
names_filials